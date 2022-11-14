import time
import openpyxl
from googletrans import Translator


def translateThisText():

    fileName = r'../HelloWorldPython/Full_Dataset.xlsx'
    linkUrlFile = openpyxl.load_workbook(fileName, data_only=True)
    sheet_obj = linkUrlFile.active
    totalReview = sheet_obj.max_row

    print("Reading Language sheet")
    linkUrlFileLang = openpyxl.load_workbook(r'../HelloWorldPython/translation.xlsx', data_only=True)
    sheetnameLang = linkUrlFileLang.get_sheet_by_name('Sheet1')

    file = open("../HelloWorldPython/translationId.txt", "r")
    readNumber = file.read()
    translationId = int(readNumber)
    file.close()
    # translationId = 2
    for i in range(translationId, totalReview):
        print('\n')
        print(i)
        cell_obj = sheet_obj.cell(row=i, column=1)
        review_Id = cell_obj.value
        cell_obj = sheet_obj.cell(row=i, column=8)
        language = cell_obj.value
        print(language)
        cell_obj = sheet_obj.cell(row=i, column=9)
        review = cell_obj.value
        print(review)
        translatedText = ""

        lengthOfReview = len(review)
        if lengthOfReview > 3900:
            translatedText = "Exceed_Length"
        else:
            if review is None:
                translatedText = "None"

            else:
                if language == "Bangla":
                    translatedText = "N/A"
                else:
                    translator = Translator()
                    translatedText = translator.translate(review, dest='bn').text
        print(translatedText)
        sheetnameLang.cell(row=i, column=1).value = review_Id
        sheetnameLang.cell(row=i, column=2).value = language
        sheetnameLang.cell(row=i, column=3).value = review
        sheetnameLang.cell(row=i, column=4).value = translatedText

        # print("Updating row_value")
        file = open("../HelloWorldPython/translationId.txt", "w")
        file.write(str(i+1))
        file.close()
        # print("Saving Language File.....")
        linkUrlFileLang.save(r'../HelloWorldPython/translation.xlsx')
        if i % 200 == 2:
            print("Waiting... you can   stop")
            time.sleep(5)
